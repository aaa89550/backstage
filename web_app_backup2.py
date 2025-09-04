#!/usr/bin/env python3
@app.route('/')
def index():
    """首頁"""
    return render_template('index.html')

@app.route('/health')
def health_check():
    """健康檢查端點"""
    return jsonify({'status': 'healthy', 'service': 'PDF Extractor'}), 200

@app.route('/upload', methods=['POST'])coding: utf-8 -*-
"""
PDF 擷取與分類程式 - Web 介面
提供網頁介面來上傳和處理 PDF 文件
"""

from flask import Flask, render_template, request, jsonify, send_file
import os
import tempfile
import pandas as pd
from advanced_pdf_extractor import AdvancedPDFExtractor
import json
from werkzeug.utils import secure_filename
import io
import zipfile

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# 初始化 PDF 擷取器
extractor = AdvancedPDFExtractor()

@app.route('/')
def index():
    """主頁面"""
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    """處理文件上傳"""
    try:
        if 'files' not in request.files:
            return jsonify({'error': '沒有選擇文件'}), 400
        
        files = request.files.getlist('files')
        if not files or files[0].filename == '':
            return jsonify({'error': '沒有選擇有效文件'}), 400
        
        results = []
        processed_files = []
        
        for file in files:
            if file and file.filename.lower().endswith('.pdf'):
                # 安全的文件名
                filename = secure_filename(file.filename)
                
                # 創建臨時文件
                with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
                    file.save(tmp_file.name)
                    
                    # 處理 PDF
                    data = extractor.process_pdf(tmp_file.name)
                    if data:
                        data['原始文件名'] = filename
                        results.append(data)
                        processed_files.append(filename)
                    
                    # 清理臨時文件
                    os.unlink(tmp_file.name)
        
        if results:
            # 保存結果到本地文件，以便備用下載
            # 確保欄位順序與前端顯示一致
            column_order = [
                'Title', '中文書名', 'Category', 'Author', 'Translator', 
                'Illustrator', 'Detail', 'Rights Sold', 'More Info', 'Tags'
            ]
            
            df = pd.DataFrame(results)
            
            # 重新排列欄位順序
            existing_columns = [col for col in column_order if col in df.columns]
            df = df[existing_columns]
            
            # 添加缺失的欄位
            for col in column_order:
                if col not in df.columns:
                    df[col] = ''
            
            # 最終確保順序正確
            df = df[column_order]
            
            try:
                # 保存 Excel
                excel_path = "/workspaces/backstage/pdf_extraction_results.xlsx"
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='PDF_Extract_Results')
                    
                    # 格式化 Excel
                    worksheet = writer.sheets['PDF_Extract_Results']
                    
                    # 調整欄位寬度
                    column_widths = {
                        'A': 50, 'B': 30, 'C': 15, 'D': 20, 'E': 20,
                        'F': 20, 'G': 80, 'H': 20, 'I': 40, 'J': 20
                    }
                    
                    for col, width in column_widths.items():
                        worksheet.column_dimensions[col].width = width
                
                # 保存 CSV
                csv_path = "/workspaces/backstage/pdf_extraction_results.csv"
                df.to_csv(csv_path, index=False, encoding='utf-8-sig')
            except Exception as save_error:
                app.logger.warning(f'保存本地文件失敗: {str(save_error)}')
            
            return jsonify({
                'success': True,
                'message': f'成功處理 {len(results)} 個 PDF 文件',
                'results': results,
                'processed_files': processed_files
            })
        else:
            return jsonify({'error': '無法從 PDF 文件中擷取內容'}), 400
            
    except Exception as e:
        return jsonify({'error': f'處理文件時發生錯誤: {str(e)}'}), 500

@app.route('/download/<format>', methods=['GET', 'POST'])
def download_results(format):
    """下載處理結果"""
    try:
        # 從請求中獲取結果數據
        if request.method == 'POST':
            results = request.json.get('results', [])
        else:
            results_json = request.args.get('data')
            if not results_json:
                return jsonify({'error': '沒有可下載的數據'}), 400
            results = json.loads(results_json)
        
        if not results:
            return jsonify({'error': '沒有可下載的數據'}), 400
        
        # 確保欄位順序與前端顯示一致
        column_order = [
            'Title', '中文書名', 'Category', 'Author', 'Translator', 
            'Illustrator', 'Detail', 'Rights Sold', 'More Info', 'Tags'
        ]
        
        # 創建 DataFrame 並確保欄位順序
        df = pd.DataFrame(results)
        
        # 重新排列欄位順序，只包含存在的欄位
        existing_columns = [col for col in column_order if col in df.columns]
        df = df[existing_columns]
        
        # 添加缺失的欄位（如果有的話）
        for col in column_order:
            if col not in df.columns:
                df[col] = ''
        
        # 最終確保順序正確
        df = df[column_order]
        
        if format.lower() == 'excel':
            # 創建 Excel 文件
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='PDF_Extract_Results')
                
                # 獲取工作表並設置格式
                worksheet = writer.sheets['PDF_Extract_Results']
                
                # 調整欄位寬度
                column_widths = {
                    'A': 50,  # Title
                    'B': 30,  # 中文書名
                    'C': 15,  # Category
                    'D': 20,  # Author
                    'E': 20,  # Translator
                    'F': 20,  # Illustrator
                    'G': 80,  # Detail
                    'H': 20,  # Rights Sold
                    'I': 40,  # More Info
                    'J': 20   # Tags
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # 設置標題行格式
                from openpyxl.styles import Font, PatternFill
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for col in range(1, len(column_order) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
            
            output.seek(0)
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='pdf_extraction_results.xlsx'
            )
            
        elif format.lower() == 'csv':
            # 創建 CSV 文件
            output = io.StringIO()
            df.to_csv(output, index=False, encoding='utf-8')
            
            # 轉換為 BytesIO
            csv_bytes = io.BytesIO()
            csv_bytes.write(output.getvalue().encode('utf-8-sig'))
            csv_bytes.seek(0)
            
            return send_file(
                csv_bytes,
                mimetype='text/csv',
                as_attachment=True,
                download_name='pdf_extraction_results.csv'
            )
        
        else:
            return jsonify({'error': '不支援的下載格式'}), 400
            
    except Exception as e:
        app.logger.error(f'下載文件時發生錯誤: {str(e)}')
        return jsonify({'error': f'下載文件時發生錯誤: {str(e)}'}), 500

@app.route('/api/process_sample')
def process_sample():
    """處理示例 PDF 文件"""
    try:
        pdf_dir = "/workspaces/backstage/PDF"
        if not os.path.exists(pdf_dir):
            return jsonify({'error': 'PDF 示例目錄不存在'}), 404
        
        pdf_files = [f for f in os.listdir(pdf_dir) if f.lower().endswith('.pdf')]
        if not pdf_files:
            return jsonify({'error': 'PDF 目錄中沒有文件'}), 404
        
        # 處理所有示例 PDF
        df = extractor.process_multiple_pdfs(pdf_dir)
        
        if not df.empty:
            results = df.to_dict('records')
            
            # 保存結果到本地文件，確保欄位順序一致
            column_order = [
                'Title', '中文書名', 'Category', 'Author', 'Translator', 
                'Illustrator', 'Detail', 'Rights Sold', 'More Info', 'Tags'
            ]
            
            # 重新排列欄位順序
            existing_columns = [col for col in column_order if col in df.columns]
            df_ordered = df[existing_columns]
            
            # 添加缺失的欄位
            for col in column_order:
                if col not in df_ordered.columns:
                    df_ordered[col] = ''
            
            # 最終確保順序正確
            df_ordered = df_ordered[column_order]
            
            try:
                # 保存 Excel
                excel_path = "/workspaces/backstage/pdf_extraction_results.xlsx"
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    df_ordered.to_excel(writer, index=False, sheet_name='PDF_Extract_Results')
                    
                    # 格式化 Excel
                    worksheet = writer.sheets['PDF_Extract_Results']
                    
                    # 調整欄位寬度
                    column_widths = {
                        'A': 50, 'B': 30, 'C': 15, 'D': 20, 'E': 20,
                        'F': 20, 'G': 80, 'H': 20, 'I': 40, 'J': 20
                    }
                    
                    for col, width in column_widths.items():
                        worksheet.column_dimensions[col].width = width
                
                # 保存 CSV
                csv_path = "/workspaces/backstage/pdf_extraction_results.csv"
                df_ordered.to_csv(csv_path, index=False, encoding='utf-8-sig')
            except Exception as save_error:
                app.logger.warning(f'保存本地文件失敗: {str(save_error)}')
            
            return jsonify({
                'success': True,
                'message': f'成功處理 {len(results)} 個示例 PDF 文件',
                'results': results,
                'processed_files': pdf_files
            })
        else:
            return jsonify({'error': '無法處理示例 PDF 文件'}), 500
            
    except Exception as e:
        return jsonify({'error': f'處理示例文件時發生錯誤: {str(e)}'}), 500

@app.route('/simple_download')
def simple_download():
    """簡單的下載頁面"""
    return '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>下載文件</title>
        <meta charset="UTF-8">
    </head>
    <body>
        <h2>下載處理結果</h2>
        <p>如果下載按鈕無法正常工作，請使用以下鏈接：</p>
        <ul>
            <li><a href="/download_file/excel">下載 Excel 文件</a></li>
            <li><a href="/download_file/csv">下載 CSV 文件</a></li>
        </ul>
        <br>
        <a href="/">返回主頁</a>
    </body>
    </html>
    '''

@app.route('/download_file/<format>')
def download_file(format):
    """直接下載最後處理的結果文件"""
    try:
        # 檢查是否有現成的結果文件
        if format.lower() == 'excel':
            file_path = "/workspaces/backstage/pdf_extraction_results.xlsx"
            if os.path.exists(file_path):
                return send_file(file_path, as_attachment=True)
        elif format.lower() == 'csv':
            file_path = "/workspaces/backstage/pdf_extraction_results.csv"
            if os.path.exists(file_path):
                return send_file(file_path, as_attachment=True)
        
        return "文件不存在，請先處理一些 PDF 文件", 404
        
    except Exception as e:
        return f"下載錯誤: {str(e)}", 500

if __name__ == '__main__':
    # 確保模板目錄存在
    template_dir = os.path.join(os.path.dirname(__file__), 'templates')
    if not os.path.exists(template_dir):
        os.makedirs(template_dir)
    
    # 從環境變量獲取端口，預設為 5000
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') != 'production'
    app.run(host='0.0.0.0', port=port, debug=debug)
